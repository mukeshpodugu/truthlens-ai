import os
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from fpdf import FPDF
from datetime import datetime

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.schemas import User, NewsArticle, Prediction, Sentiment, Emotion, CredibilityScore, SavedReport, SavedReportCreate, SavedReportResponse, ActivityLog

router = APIRouter(prefix="/reports", tags=["Saved Reports & Exports"])

@router.post("/save", response_model=SavedReportResponse)
def save_report(report_in: SavedReportCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Verify article exists
    article = db.query(NewsArticle).filter(NewsArticle.id == report_in.article_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
        
    db_report = SavedReport(
        user_id=current_user.id,
        article_id=report_in.article_id,
        title=report_in.title,
        notes=report_in.notes
    )
    db.add(db_report)
    db.commit()
    db.refresh(db_report)
    
    log = ActivityLog(user_id=current_user.id, action=f"Saved report bookmark: {report_in.title}")
    db.add(log)
    db.commit()
    
    return db_report

@router.get("/list", response_model=list[SavedReportResponse])
def list_saved_reports(
    search: str = Query(None, description="Search by title or notes"),
    category: str = Query(None, description="Filter by news category"),
    sentiment: str = Query(None, description="Filter by sentiment label"),
    min_credibility: int = Query(None, description="Filter by minimum credibility score"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(SavedReport).join(NewsArticle, SavedReport.article_id == NewsArticle.id).filter(SavedReport.user_id == current_user.id)
    
    if search:
        query = query.filter(
            (SavedReport.title.ilike(f"%{search}%")) | (SavedReport.notes.ilike(f"%{search}%"))
        )
        
    if category:
        from app.models.schemas import Category
        query = query.join(Category, NewsArticle.category_id == Category.id).filter(Category.name.ilike(category))
        
    if sentiment:
        query = query.join(Sentiment, Sentiment.article_id == NewsArticle.id).filter(Sentiment.label == sentiment.lower())
        
    if min_credibility is not None:
        query = query.join(CredibilityScore, CredibilityScore.article_id == NewsArticle.id).filter(CredibilityScore.score >= min_credibility)
        
    return query.order_by(SavedReport.created_at.desc()).all()

@router.delete("/delete/{report_id}")
def delete_saved_report(report_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    report = db.query(SavedReport).filter(SavedReport.id == report_id, SavedReport.user_id == current_user.id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Saved report not found")
        
    title = report.title
    db.delete(report)
    db.commit()
    
    log = ActivityLog(user_id=current_user.id, action=f"Deleted saved report: {title}")
    db.add(log)
    db.commit()
    
    return {"status": "success", "message": f"Deleted report '{title}' bookmark successfully"}

class TruthLensPDF(FPDF):
    def header(self):
        # Top banner styling
        self.set_fill_color(30, 27, 75)  # dark indigo
        self.rect(0, 0, 210, 35, 'F')
        
        self.set_text_color(255, 255, 255)
        self.set_font('helvetica', 'B', 20)
        self.cell(0, 15, 'TRUTHLENS AI - MEDIA ANALYSIS REPORT', border=0, ln=1, align='C')
        self.set_font('helvetica', 'I', 10)
        self.cell(0, 5, 'Intelligent Misinformation & Credibility Analysis Platform', border=0, ln=1, align='C')
        self.ln(15)
        
    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-25)
        self.set_font('helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, 'TruthLens AI Platform developed by PODUGU MUKESH (mukeshpodugu123@gmail.com | 8143999463).', border=0, ln=1, align='C')
        self.cell(0, 5, f'Page {self.page_no()}/{{nb}} - Generated on {datetime.now().strftime("%Y-%m-%d %H:%M")}', border=0, align='C')

@router.get("/export/{article_id}/pdf")
def export_pdf_report(article_id: int, db: Session = Depends(get_db)):
    article = db.query(NewsArticle).filter(NewsArticle.id == article_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
        
    pred = article.predictions[0] if article.predictions else None
    cred = article.credibility_scores[0] if article.credibility_scores else None
    sent = article.sentiments[0] if article.sentiments else None
    emot = article.emotions[0] if article.emotions else None
    
    pdf = TruthLensPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('helvetica', '', 11)
    pdf.set_text_color(30, 41, 59)
    
    # 1. Metadata Section
    pdf.set_font('helvetica', 'B', 14)
    pdf.cell(0, 8, f"Document: {article.title[:65]}...", border=0, ln=1)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(0, 6, f"Source URL: {article.source_url or 'N/A'}  |  Publisher: {article.publisher or 'N/A'}", border=0, ln=1)
    pdf.cell(0, 6, f"Analyzed on: {article.created_at.strftime('%Y-%m-%d %H:%M:%S')}  |  Model: {pred.model_name if pred else 'N/A'}", border=0, ln=1)
    pdf.ln(5)
    
    # 2. Main Analytics Table
    pdf.set_fill_color(241, 245, 249)
    pdf.set_font('helvetica', 'B', 11)
    pdf.cell(50, 10, 'Metric', border=1, fill=True)
    pdf.cell(80, 10, 'Analysis Value', border=1, fill=True)
    pdf.cell(60, 10, 'Assessment Scale', border=1, fill=True, ln=1)
    
    pdf.set_font('helvetica', '', 10)
    pdf.cell(50, 8, 'Credibility Score', border=1)
    pdf.cell(80, 8, f"{cred.score if cred else 'N/A'}/100 - {cred.status if cred else 'N/A'}", border=1)
    pdf.cell(60, 8, '80+ Credible, <40 Unreliable', border=1, ln=1)
    
    pdf.cell(50, 8, 'Fake News Class', border=1)
    pdf.cell(80, 8, f"{pred.label if pred else 'N/A'} ({round(pred.confidence_score * 100, 2) if pred else 0.0}% confidence)", border=1)
    pdf.cell(60, 8, 'Real, Fake, Misleading, Satire', border=1, ln=1)
    
    pdf.cell(50, 8, 'Sentiment Label', border=1)
    pdf.cell(80, 8, f"{sent.label.upper() if sent else 'N/A'} (Pos: {sent.pos_score if sent else 0.0}, Neg: {sent.neg_score if sent else 0.0})", border=1)
    pdf.cell(60, 8, 'Positive, Negative, Neutral', border=1, ln=1)
    
    pdf.cell(50, 8, 'Dominant Emotion', border=1)
    pdf.cell(80, 8, f"{emot.dominant_emotion.upper() if emot else 'N/A'}", border=1)
    pdf.cell(60, 8, 'Anger, Fear, Joy, Surprise, Sad', border=1, ln=1)
    
    pdf.ln(8)
    
    # 3. AI Explanation & Suspicious Phrases
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 8, "Explainable AI Insights:", border=0, ln=1)
    pdf.set_font('helvetica', '', 10)
    
    explanation_text = "The analysis indicates that the article's wording contains "
    if pred and pred.label in ["Fake News", "Misleading Content", "Satire"]:
        explanation_text += "exaggerated clickbait modifiers, unverified conspiracy tropes, and emotionally polarizing words that are statistically common in fake news datasets."
    else:
        explanation_text += "balanced reporting styles, low emotional bias vocabulary, and syntactic alignments standard to factual journalism reviews."
        
    pdf.multi_cell(0, 6, explanation_text, border=0)
    pdf.ln(5)
    
    # 4. Original Text snippet
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 8, "Parsed Content Snippet:", border=0, ln=1)
    pdf.set_font('helvetica', 'I', 9)
    snippet = article.text[:400] + "..." if len(article.text) > 400 else article.text
    pdf.multi_cell(0, 5, snippet, border=1)
    
    # Output PDF byte buffer
    pdf_bytes = pdf.output(dest='S')
    
    buffer = BytesIO(pdf_bytes.encode('latin1') if isinstance(pdf_bytes, str) else pdf_bytes)
    return StreamingResponse(
        buffer, 
        media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename=truthlens_report_{article_id}.pdf"}
    )

@router.get("/export/{article_id}/excel")
def export_excel_report(article_id: int, db: Session = Depends(get_db)):
    article = db.query(NewsArticle).filter(NewsArticle.id == article_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
        
    pred = article.predictions[0] if article.predictions else None
    cred = article.credibility_scores[0] if article.credibility_scores else None
    sent = article.sentiments[0] if article.sentiments else None
    emot = article.emotions[0] if article.emotions else None
    
    # Generate Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TruthLens Analysis"
    
    # Formatting styles
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 55
    
    ws['A1'] = "TruthLens AI Analysis Report"
    ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
    
    data = [
        ("Article ID", article.id),
        ("Title", article.title),
        ("Source URL", article.source_url or "N/A"),
        ("Publisher", article.publisher or "N/A"),
        ("Analysis Timestamp", article.created_at.strftime('%Y-%m-%d %H:%M:%S')),
        ("Model Classifier Used", pred.model_name if pred else "N/A"),
        ("Classification Verdict", pred.label if pred else "N/A"),
        ("Model Confidence Score", pred.confidence_score if pred else 0.0),
        ("Credibility Score (0-100)", cred.score if cred else 0),
        ("Credibility Assessment", cred.status if cred else "N/A"),
        ("Dominant Sentiment", sent.label if sent else "N/A"),
        ("Sentiment Positive Ratio", sent.pos_score if sent else 0.0),
        ("Sentiment Negative Ratio", sent.neg_score if sent else 0.0),
        ("Dominant Emotion detected", emot.dominant_emotion if emot else "N/A"),
        ("Developer Name", "PODUGU MUKESH"),
        ("Developer Email", "mukeshpodugu123@gmail.com")
    ]
    
    for row_idx, (metric, val) in enumerate(data, start=3):
        ws.cell(row=row_idx, column=1, value=metric).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row_idx, column=2, value=val)
        
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=truthlens_metrics_{article_id}.xlsx"}
    )
